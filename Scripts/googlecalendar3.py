import pickle
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from datetime import datetime, timedelta
import pytz
import openpyxl, os
from docx import Document

from Google import Create_Service
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import numpy as np

scopes = ['https://www.googleapis.com/auth/calendar', 'https://www.googleapis.com/auth/calendar.readonly', 'https://www.googleapis.com/auth/calendar.events', 'https://www.googleapis.com/auth/calendar.events.readonly']
flow = InstalledAppFlow.from_client_secrets_file("client_secret_GoogleCloudDemo.json", scopes=scopes)
credentials = flow.run_console()

pickle.dump(credentials, open("token.pkl", "wb"))
credentials = pickle.load(open("token.pkl", "rb"))
service = build("calendar", "v3", credentials=credentials)

result = service.calendarList().list().execute()
print(result['items'][0])
calendar_id = result['items'][0]['id']

def format_date(date):
    x = date.split('/')
    return datetime(2021, int(x[0]), int(x[1]), 13, 0, 0)

# Load excel file
xl = openpyxl.load_workbook('C:/Users/Karan/Desktop/course.xlsx')

#Authorize the API
scope = ['https://www.googleapis.com/auth/drive','https://www.googleapis.com/auth/drive.file']
file_name = 'client_service.json'
creds = ServiceAccountCredentials.from_json_keyfile_name(file_name,scope)
client = gspread.authorize(creds)

gauth = GoogleAuth()
gauth.LocalWebserverAuth()

drive = GoogleDrive(gauth)

#Fetch the sheet
worksheet = client.open('IT212_Schedule').sheet1

sheet = xl['Sheet1']

#columns and index
letters = ["A","B","C","E","F","G","H","I"]

r = 0               # Keeps track of rows
w = 2               # Keep track of biweekly count

remainder = (sheet.max_row - 1) % 2

maxRow = (((sheet.max_row - 1) // 2) + remainder)

array = np.array(worksheet.get_all_values())

arr = []

for x in array:
    for y in x:
        arr.append(y)

def remove():
    for element in range(9):
        arr.pop(0)

remove()


# Parse through worksheet to manipulate cells
for row in range(maxRow):
    # if value in certain cells is not none then create event for it
    c1 = worksheet.acell('A' + str(w)).value
    c3 = worksheet.acell('H' + str(w)).value
    c4 = worksheet.acell('I' + str(w)).value
    c7 = worksheet.acell('E' + str(w)).value
    if c1 is not None:
        start_time = format_date(c1)
        end_time = start_time + timedelta(hours=2)
        timezone = 'America/New_York'
        event = {
            'summary': 'Class',
            'location': 'EnGeo 2209',
            'description': c7,
            'start': {
                'dateTime': start_time.strftime("%Y-%m-%dT%H:%M:%S"),
                'timeZone': timezone,
            },
            'end': {
                'dateTime': end_time.strftime("%Y-%m-%dT%H:%M:%S"),
                'timeZone': timezone,
            },
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'email', 'minutes': 24 * 60},
                    {'method': 'popup', 'minutes': 10},
                ],
            },
        }
        service.events().insert(calendarId=calendar_id, body=event).execute()
    if c3 is not None:
        start_time = format_date(c1)
        end_time = start_time + timedelta(hours=2)
        timezone = 'America/New_York'
        event = {
            'summary': 'HW ' + c3 + ' due',
            'location': 'EnGeo 2209',
            'description': 'Homework',
            'start': {
                'dateTime': start_time.strftime("%Y-%m-%dT%H:%M:%S"),
                'timeZone': timezone,
            },
            'end': {
                'dateTime': end_time.strftime("%Y-%m-%dT%H:%M:%S"),
                'timeZone': timezone,
            },
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'email', 'minutes': 24 * 60},
                    {'method': 'popup', 'minutes': 10},
                ],
            },
        }
        service.events().insert(calendarId=calendar_id, body=event).execute()
    if c4 is not None:
        start_time = format_date(c1)
        end_time = start_time + timedelta(hours=2)
        timezone = 'America/New_York'
        event = {
            'summary': 'Lab ' + c4 + ' due',
            'location': 'EnGeo 2209',
            'description': c7,
            'start': {
                'dateTime': start_time.strftime("%Y-%m-%dT%H:%M:%S"),
                'timeZone': timezone,
            },
            'end': {
                'dateTime': end_time.strftime("%Y-%m-%dT%H:%M:%S"),
                'timeZone': timezone,
            },
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'email', 'minutes': 24 * 60},
                    {'method': 'popup', 'minutes': 10},
                ],
            },
        }
        service.events().insert(calendarId=calendar_id, body=event).execute()
    w += 1
